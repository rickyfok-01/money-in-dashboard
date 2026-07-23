"""Convert the screenshot table (trustee × month: Pay AMT / Avail AMT / ALLOC%)
into a formatted Excel workbook. One-off; not part of the data pipeline."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MONTHS = ["Feb 2026 contribution", "Mar 2026 contribution",
          "Apr 2026 contribution", "May 2026 contribution"]

# (trustee, [(pay, avail, alloc_frac), ...]) — None = no figure in source
ROWS = [
    ("EA",        [(377.9, 5.8, .985), (342.6, 4.6, .987), (360.0, 28.4, .921), (341.6, 9.5, .972)]),
    ("PRIN",      [(330.9, 4.8, .986), (326.2, 5.6, .983), (321.1, 5.6, .983), (314.8, 9.7, .969)]),
    ("BOCI",      [(788.4, 10.6, .987),(785.7, 13.4, .983),(790.0, 12.7, .984),(761.9, 19.8, .974)]),
    ("BCT1",      [(208.7, 2.0, .990), (213.0, 0.8, .996), (200.5, 1.4, .993), (204.9, 8.3, .959)]),
    ("BCT2",      [(480.2, 3.9, .992), (476.4, 4.4, .991), (448.6, 4.8, .989), (440.0, 10.6, .978)]),
    ("AIA",       [(763.4, 8.8, .989), (745.7, 14.8, .980),(690.7, 10.7, .985),(711.1, 18.3, .974)]),
    ("SUN",       [(1023.6, 7.5, .993),(982.8, 7.3, .993), (987.4, 7.7, .992), (967.7, 11.5, .988)]),
    ("MAN",       [(2127.5, 14.3, .993),(1906.4, 29.1, .985),(1802.2, 22.8, .987),(1801.1, 60.5, .966)]),
    ("HSBC 1",    [(524.7, 3.0, .994), (516.8, 2.4, .995), (496.9, 2.5, .995), (498.6, 4.4, .991)]),
    ("HSBC 2",    [(1936.3, 30.2, .984),(1889.0, 10.0, .995),(1734.7, 10.9, .994),(1715.8, 15.4, .991)]),
    ("BCT (IC)",  [(None, None, None), (49.8, 9.3, .814),  (47.4, 9.0, .811),  (49.1, 10.7, .781)]),
    ("BEA (IS)",  [(None, None, None), (None, None, None), (150.3, 30.6, .797),(153.9, 33.9, .708)]),
]
GRAND_TOTAL = [(8561.7, 90.9, .989), (8234.2, 101.8, .988), (8029.8, 147.0, .982), (7960.5, 212.6, .973)]
LAST_WEEK   = [(8559.7, 93.5, .989), (8231.0, 106.9, .987), (8024.6, 163.1, .980), (7960.2, 306.9, .961)]

# --- styles -----------------------------------------------------------------
INK = "1A1A1A"
hdr_fill   = PatternFill("solid", fgColor="1F3A5F")
sub_fill   = PatternFill("solid", fgColor="E8EDF3")
tot_fill   = PatternFill("solid", fgColor="D9E2F0")
lv_fill    = PatternFill("solid", fgColor="F0E9D6")
white_bold = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
bold       = Font(name="Calibri", bold=True, color=INK)
italic     = Font(name="Calibri", italic=True, color="6B6B6B", size=9)
center     = Alignment(horizontal="center", vertical="center", wrap_text=True)
left       = Alignment(horizontal="left", vertical="center")
right      = Alignment(horizontal="right", vertical="center")
thin = Side(style="thin", color="BFC9D4")
med  = Side(style="medium", color=INK)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
ws = wb.active
ws.title = "Allocation"
ws.sheet_view.showGridLines = False

NCOLS = 1 + 4 * 3  # trustee + 4 months × 3 cols = 13

# Row 1 — title
ws.cell(1, 1, "Contribution Allocation by Trustee (Pay AMT / Avail AMT / ALLOC %)").font = \
    Font(name="Calibri", bold=True, size=14, color=INK)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
ws.cell(1, 1).alignment = left

# Row 2 — month group headers (merged across 3 cols each)
ws.cell(2, 1, "Trustee")
ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
for i, m in enumerate(MONTHS):
    c0 = 2 + i * 3
    ws.merge_cells(start_row=2, start_column=c0, end_row=2, end_column=c0 + 2)
    cell = ws.cell(2, c0, m)

# Row 3 — sub-headers
subs = ["Pay AMT", "Avail AMT", "ALLOC(%)"]
for i in range(4):
    for j, s in enumerate(subs):
        ws.cell(3, 2 + i * 3 + j, s)

# style header block
for r in (2, 3):
    for c in range(1, NCOLS + 1):
        cell = ws.cell(r, c)
        cell.fill = hdr_fill
        cell.font = white_bold
        cell.alignment = center
        cell.border = border

# Row 4 — "Last version (30 Jun)" annotation, merged per month group
for i in range(4):
    c0 = 2 + i * 3
    ws.merge_cells(start_row=4, start_column=c0, end_row=4, end_column=c0 + 2)
    cell = ws.cell(4, c0, "Last version (30 Jun)")
    cell.fill = lv_fill
    cell.font = italic
    cell.alignment = center
    for c in range(c0, c0 + 3):
        ws.cell(4, c).border = border
ws.cell(4, 1).border = border

# Data rows
r = 5
for name, months in ROWS:
    ws.cell(r, 1, name).font = bold
    ws.cell(r, 1).alignment = left
    for i, (pay, avail, alloc) in enumerate(months):
        base = 2 + i * 3
        if pay is not None:
            ws.cell(r, base, pay).number_format = "#,##0.0"
        if avail is not None:
            ws.cell(r, base + 1, avail).number_format = "#,##0.0"
        if alloc is not None:
            ws.cell(r, base + 2, alloc).number_format = "0.0%"
        else:
            ws.cell(r, base + 2, "/").alignment = center
        for off in range(3):
            ws.cell(r, base + off).alignment = right if off < 2 else center
    for c in range(1, NCOLS + 1):
        ws.cell(r, c).border = border
    r += 1

# Grand Total + Last week total (with medium top border)
def write_total(row, label, data, fill):
    ws.cell(row, 1, label).font = bold
    ws.cell(row, 1).alignment = left
    for i, (pay, avail, alloc) in enumerate(data):
        base = 2 + i * 3
        ws.cell(row, base, pay).number_format = "#,##0.0"
        ws.cell(row, base + 1, avail).number_format = "#,##0.0"
        ws.cell(row, base + 2, alloc).number_format = "0.0%"
        for off in range(3):
            ws.cell(row, base + off).alignment = right if off < 2 else center
            ws.cell(row, base + off).font = bold
    for c in range(1, NCOLS + 1):
        cell = ws.cell(row, c)
        cell.fill = fill
        cell.border = Border(left=thin, right=thin,
                             top=med if row == r else thin, bottom=thin)

write_total(r, "Grand Total", GRAND_TOTAL, tot_fill); r += 1
write_total(r, "Last week total", LAST_WEEK, tot_fill)

# Column widths + freeze
ws.column_dimensions["A"].width = 16
for i in range(2, NCOLS + 1):
    ws.column_dimensions[get_column_letter(i)].width = 11
ws.freeze_panes = "B5"

out = "contribution-allocation.xlsx"
wb.save(out)
print("wrote", out)
