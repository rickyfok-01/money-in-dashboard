#!/usr/bin/env python3
"""
build_data.py — Contribution snapshot CSV -> data.js

Reads every data/con-bill-6mon-YYYYMMDD.csv (Query 2 of contribution.sql),
coerces every measure to int (the 20260708 export ships floats like 9.0;
20260707 ships ints — normalize both), keeps only the latest 6 distinct
YEAR_MONTHs across all snapshots, and writes a single `const DATA = {...}`
to data.js.

Also reads the payment dataset data/con-pym-6mon-YYYYMMDD.csv (PAY_AMT /
AVAIL_AMOUNT by scheme x month) into DATA.pym, and the code->name lookup
data/constant-scheme-info.xlsx into DATA.names. Bill data and a/b flags are
unaffected by these additions.

Re-run whenever the CSVs change. Never hand-edit data.js.

    python scripts/build_data.py
"""
from __future__ import annotations
import csv
import glob
import json
import os
import re
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DATA_DIR = os.path.join(ROOT, "data")
OUT = os.path.join(ROOT, "data.js")

GLOB = os.path.join(DATA_DIR, "con-bill-6mon-*.csv")
PYM_GLOB = os.path.join(DATA_DIR, "con-pym-6mon-*.csv")
AO_AGING_GLOB = os.path.join(DATA_DIR, "con-pym-ao-aging-*.csv")
DDI30_GLOB = os.path.join(DATA_DIR, "ddi-30day-*.csv")
DDI_AGING_GLOB = os.path.join(DATA_DIR, "ddi-aging-*.csv")
DDA30_GLOB = os.path.join(DATA_DIR, "dda-30day-*.csv")
DDA_AGING_GLOB = os.path.join(DATA_DIR, "dda-aging-*.csv")
NAMES_XLSX = os.path.join(DATA_DIR, "constant-scheme-info.xlsx")

# Canonical status lifecycle order (forward path first, then edge/terminal).
STATUS_ORDER = [
    "OPEN", "PARTIAL_SUBMIT", "SUBMITTED", "APPROVED",
    "PARTIAL_PAID", "FULLY_PAID", "CLOSED",
    "OVERPAID", "REFUND_OVERPAID", "WAIVED",
]
# Contribution-mode display order.
MODE_ORDER = ["REGULAR", "LUMP_SUM", "SURCHARGE"]
# Channel order is structural (identity) — fixed.
CHANNELS = ["DDE", "BATCH", "PORTAL", "BULKUPLOAD", "OTHER"]
FREQ_DISPLAY_ORDER = [
    "DAILY", "WEEKLY", "FORTNIGHTLY", "SEMI_MONTHLY", "MONTHLY",
    "BI_MONTHLY", "QUARTERLY", "SEMI_ANNUALLY", "ANNUALLY",
]
UNSET = "(unset)"

KEEP_MONTHS = 6  # latest N distinct YEAR_MONTHs


def snap_from_name(path: str) -> str:
    """con-bill-6mon-20260708.csv -> '20260708'."""
    m = re.search(r"(\d{8})\.csv$", os.path.basename(path))
    return m.group(1) if m else ""


# Statuses that count as "ER submitted contribution data" (A) — i.e. the bill
# has been submitted by the employer (tagged/approved/paid) per the spec.
A_STATUSES = {"PARTIAL_SUBMIT", "SUBMITTED", "APPROVED", "PARTIAL_PAID", "FULLY_PAID"}
# Statuses that count as "Pending Tagging" (B) — submitted but not yet paid.
B_STATUSES = {"PARTIAL_SUBMIT", "SUBMITTED", "APPROVED"}


def to_int(v) -> int:
    """Coerce '9.0' / '9' / '' -> 9 / 9 / 0."""
    if v is None:
        return 0
    s = str(v).strip()
    if s == "":
        return 0
    try:
        return int(round(float(s)))
    except ValueError:
        return 0


def to_float(v) -> float:
    """Coerce '72900929.46' / '' -> 72900929.46 / 0.0 (rounded to cents)."""
    if v is None:
        return 0.0
    s = str(v).strip()
    if s == "":
        return 0.0
    try:
        return round(float(s), 2)
    except ValueError:
        return 0.0


def _blank_row(r) -> bool:
    """True when a DictReader row has no TR_CODE and no SCHEME_CODE (blank
    source line, all-empty-comma line, or a garbled non-CSV line parsed under
    a wrong header). Such rows are artifacts: `snap_file` backfills the
    snapshot from the filename, so the `if not snap` guard alone does not
    catch them and they would otherwise surface as `{tr:'',sc:'',...}` rows
    in the output. Applied in every reader loop."""
    tr = (r.get("TR_CODE") or "").strip()
    sc = (r.get("SCHEME_CODE") or "").strip()
    return not tr and not sc


def freq_label(v: str) -> str:
    v = (v or "").strip()
    return v if v else UNSET


def read_pym():
    """Read con-pym-6mon-*.csv -> (rows, snapshots, months).

    Each row is one combination of snapshot x trustee x scheme x pay channel x
    tag status x pay method x contribution month, with measures PAYMENT_COUNT
    (int) and PAY_AMT / AVAIL_AMOUNT (float HKD, rounded to cents).
    """
    rows, snaps, months = [], set(), set()
    for p in sorted(glob.glob(PYM_GLOB)):
        snap_file = snap_from_name(p)
        with open(p, newline="", encoding="utf-8-sig") as fh:
            for r in csv.DictReader(fh):
                snap = (r.get("SNAPSHOT_DATE") or snap_file or "").strip()
                if not snap:
                    continue
                if _blank_row(r):
                    continue
                row = {
                    "s": snap,
                    "tr": (r.get("TR_CODE") or "").strip(),
                    "sc": (r.get("SCHEME_CODE") or "").strip(),
                    "chan": (r.get("AV_PAY_CHANNEL_CODE") or "").strip(),
                    "tag": (r.get("AV_TAG_STATUS_CODE") or "").strip(),
                    "pm": (r.get("PAY_METHOD_CODE") or "").strip(),
                    "ym": (r.get("MONTH") or "").strip(),
                    "pc": to_int(r.get("PAYMENT_COUNT")),
                    "pay": to_float(r.get("PAY_AMT")),
                    "avail": to_float(r.get("AVAIL_AMOUNT")),
                }
                rows.append(row)
                snaps.add(row["s"])
                months.add(row["ym"])
    return rows, snaps, months


def read_ao_aging():
    """Read con-pym-ao-aging-*.csv -> list of rows.
    Each row: snapshot, tr, sc, chan, tag, pm, total, day_00_06..day_31_more.
    """
    rows, snaps = [], set()
    for p in sorted(glob.glob(AO_AGING_GLOB)):
        snap_file = snap_from_name(p)
        with open(p, newline="", encoding="utf-8-sig") as fh:
            for r in csv.DictReader(fh):
                snap = (r.get("SNAPSHOT_DATE") or snap_file or "").strip()
                if not snap:
                    continue
                if _blank_row(r):
                    continue
                row = {
                    "s": snap,
                    "tr": (r.get("TR_CODE") or "").strip(),
                    "sc": (r.get("SCHEME_CODE") or "").strip(),
                    "chan": (r.get("AV_PAY_CHANNEL_CODE") or "").strip(),
                    "tag": (r.get("AV_TAG_STATUS_CODE") or "").strip(),
                    "pm": (r.get("PAY_METHOD_CODE") or "").strip(),
                    "total": to_int(r.get("TOTAL")),
                    "d00_06": to_int(r.get("DAY_00_06")),
                    "d07_14": to_int(r.get("DAY_07_14")),
                    "d15_21": to_int(r.get("DAY_15_21")),
                    "d22_30": to_int(r.get("DAY_22_30")),
                    "d31": to_int(r.get("DAY_31_MORE")),
                }
                rows.append(row)
                snaps.add(row["s"])
    return rows, sorted(snaps)


def read_ddi():
    """Read ddi-30day-*.csv and ddi-aging-*.csv -> (last30, aging) row lists."""
    last30, aging = [], []
    snaps30, snapsAging = set(), set()

    for p in sorted(glob.glob(DDI30_GLOB)):
        snap_file = snap_from_name(p)
        with open(p, newline="", encoding="utf-8-sig") as fh:
            for r in csv.DictReader(fh):
                snap = (r.get("SNAPSHOT_DATE") or snap_file or "").strip()
                if not snap:
                    continue
                if _blank_row(r):
                    continue
                row = {
                    "s": snap,
                    "tr": (r.get("TR_CODE") or "").strip(),
                    "sc": (r.get("SCHEME_CODE") or "").strip(),
                    "at": (r.get("SHORT_CODE") or "").strip(),
                    "date": (r.get("DDI_REQUEST_DATE") or "").strip(),
                    "total": to_int(r.get("COUNT")),
                    "submitted": to_int(r.get("SUBMITTED_TO_BANK")),
                    "success": to_int(r.get("SUCCESS")),
                    "rejected": to_int(r.get("REJECTED")),
                }
                last30.append(row)
                snaps30.add(row["s"])

    for p in sorted(glob.glob(DDI_AGING_GLOB)):
        snap_file = snap_from_name(p)
        with open(p, newline="", encoding="utf-8-sig") as fh:
            for r in csv.DictReader(fh):
                snap = (r.get("SNAPSHOT_DATE") or snap_file or "").strip()
                if not snap:
                    continue
                if _blank_row(r):
                    continue
                row = {
                    "s": snap,
                    "tr": (r.get("TR_CODE") or "").strip(),
                    "sc": (r.get("SCHEME_CODE") or "").strip(),
                    "at": (r.get("SHORT_CODE") or "").strip(),
                    "total": to_int(r.get("TOTAL")),
                    "d00_06": to_int(r.get("DAY_00_06")),
                    "d07_14": to_int(r.get("DAY_07_14")),
                    "d15_21": to_int(r.get("DAY_15_21")),
                    "d22_30": to_int(r.get("DAY_22_30")),
                    "d31": to_int(r.get("DAY_31_MORE")),
                }
                aging.append(row)
                snapsAging.add(row["s"])

    return last30, aging, sorted(snaps30), sorted(snapsAging)


def read_dda():
    """Read dda-30day-*.csv and dda-aging-*.csv -> (last30, aging) row lists."""
    last30, aging = [], []
    snaps30, snapsAging = set(), set()

    for p in sorted(glob.glob(DDA30_GLOB)):
        snap_file = snap_from_name(p)
        with open(p, newline="", encoding="utf-8-sig") as fh:
            for r in csv.DictReader(fh):
                snap = (r.get("SNAPSHOT_DATE") or snap_file or "").strip()
                if not snap:
                    continue
                if _blank_row(r):
                    continue
                row = {
                    "s": snap,
                    "tr": (r.get("TR_CODE") or "").strip(),
                    "sc": (r.get("SCHEME_CODE") or "").strip(),
                    "at": (r.get("SHORT_CODE") or "").strip(),
                    "total": to_int(r.get("TOTAL")),
                    "submitted_pig": to_int(r.get("SUBMITTED_TO_PIG")),
                    "submitted_bank": to_int(r.get("SUBMITTED_TO_BANK")),
                    "active": to_int(r.get("ACTIVE")),
                    "inactive": to_int(r.get("INACTIVE")),
                    "rejected": to_int(r.get("REJECTED")),
                    "suspend": to_int(r.get("SUSPEND")),
                }
                last30.append(row)
                snaps30.add(row["s"])

    for p in sorted(glob.glob(DDA_AGING_GLOB)):
        snap_file = snap_from_name(p)
        with open(p, newline="", encoding="utf-8-sig") as fh:
            for r in csv.DictReader(fh):
                snap = (r.get("SNAPSHOT_DATE") or snap_file or "").strip()
                if not snap:
                    continue
                if _blank_row(r):
                    continue
                row = {
                    "s": snap,
                    "tr": (r.get("TR_CODE") or "").strip(),
                    "sc": (r.get("SCHEME_CODE") or "").strip(),
                    "at": (r.get("SHORT_CODE") or "").strip(),
                    "total": to_int(r.get("TOTAL")),
                    "d00_06": to_int(r.get("DAY_00_06")),
                    "d07_14": to_int(r.get("DAY_07_14")),
                    "d15_21": to_int(r.get("DAY_15_21")),
                    "d22_30": to_int(r.get("DAY_22_30")),
                    "d31": to_int(r.get("DAY_31_MORE")),
                }
                aging.append(row)
                snapsAging.add(row["s"])

    return last30, aging, sorted(snaps30), sorted(snapsAging)


def read_names():
    """constant-scheme-info.xlsx -> {scheme:{code:name}, trustee:{code:name}}.

    Returns empty maps if openpyxl is missing or the file is absent (the
    dashboard degrades to showing raw codes).
    """
    out = {"scheme": {}, "trustee": {}}
    if not os.path.exists(NAMES_XLSX):
        return out
    try:
        import openpyxl
    except ImportError:
        return out
    wb = openpyxl.load_workbook(NAMES_XLSX, data_only=True)
    ws = wb["Export Worksheet"] if "Export Worksheet" in wb.sheetnames else wb[wb.sheetnames[0]]
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        tr = str(row[0]).strip()
        if tr.upper() == "TR_CODE":
            continue  # header row
        tr_name = (row[1] is not None and str(row[1]).strip()) or ""
        sc = (len(row) > 2 and row[2] is not None and str(row[2]).strip()) or ""
        sc_name = (len(row) > 3 and row[3] is not None and str(row[3]).strip()) or ""
        if tr and tr_name:
            out["trustee"].setdefault(tr, tr_name)
        if sc and sc_name:
            out["scheme"].setdefault(sc, sc_name)
    return out


def main() -> int:
    paths = sorted(glob.glob(GLOB))
    if not paths:
        print(f"[build_data] no CSVs matching {GLOB}", file=sys.stderr)
        return 1

    raw_rows = []
    snapshots = set()
    months = set()
    schemes = set()
    trustees = set()
    statuses = set()
    modes = set()
    freqs = set()
    acct_types = set()

    for p in paths:
        snap_file = snap_from_name(p)
        with open(p, newline="", encoding="utf-8-sig") as fh:
            reader = csv.DictReader(fh)
            for r in reader:
                # Authoritative snapshot: the column, falling back to filename.
                snap = (r.get("SNAPSHOT_DATE") or snap_file or "").strip()
                if not snap:
                    continue
                if _blank_row(r):
                    continue
                ym = (r.get("YEAR_MONTH") or "").strip()
                sc = (r.get("SCHEME_CODE") or "").strip()
                tr = (r.get("TR_CODE") or "").strip()
                st = (r.get("AV_STATUS_CODE") or "").strip()
                bm = (r.get("AV_BILL_CONTR_MODE") or "").strip()
                fq = freq_label(r.get("AV_FREQ_TYPE") or "")
                at = (r.get("SHORT_CODE") or "").strip()

                row = {
                    "s": snap, "tr": tr, "sc": sc, "st": st,
                    "bm": bm, "fq": fq, "at": at, "ym": ym,
                    "bill":   to_int(r.get("BILL_COUNT")),
                    "ontime": to_int(r.get("ONTIME_SUBMIT_COUNT")),
                    "total":  to_int(r.get("TOTAL_SUBMIT_COUNT")),
                    "dde":    to_int(r.get("DDE_SUBMIT_COUNT")),
                    "batch":  to_int(r.get("BATCH_SUBMIT_COUNT")),
                    "portal": to_int(r.get("PORTAL_SUBMIT_COUNT")),
                    "bulk":   to_int(r.get("BULKUPLOAD_SUBMIT_COUNT")),
                    "other":  to_int(r.get("OTHER_SUBMIT_COUNT")),
                    # a = ER submitted contribution data (per-row indicator; 1 if
                    #     this row's status counts as A, else 0). b = pending
                    #     tagging (per-row indicator; 1 if status is still
                    #     pre-payment). The renderer sums them per group so a row
                    #     with bill=N contributes N to a/b if the status matches.
                    "a": to_int(r.get("BILL_COUNT")) if st in A_STATUSES else 0,
                    "b": to_int(r.get("BILL_COUNT")) if st in B_STATUSES else 0,
                }
                raw_rows.append(row)
                snapshots.add(snap)
                months.add(ym)
                schemes.add(sc)
                trustees.add(tr)
                statuses.add(st)
                modes.add(bm)
                freqs.add(fq)
                acct_types.add(at)

    # Latest KEEP_MONTHS distinct YEAR_MONTHs (string sort works for YYYY-MM).
    months_sorted = sorted(m for m in months if m)
    keep = set(months_sorted[-KEEP_MONTHS:])
    rows = [r for r in raw_rows if r["ym"] in keep]

    # Payment (pym) dataset — separate grain, same latest-6-month window.
    pym_rows_raw, pym_snaps, pym_months = read_pym()
    pym_rows = [r for r in pym_rows_raw if r["ym"] in keep]
    pym = {
        "snapshots": sorted(s for s in pym_snaps if s),
        "months": sorted(m for m in pym_months if m and m in keep),
        "rows": pym_rows,
    }

    # AO aging dataset.
    ao_aging_rows, ao_aging_snaps = read_ao_aging()

    # DDI datasets.
    ddi30_rows, ddi_aging_rows, ddi30_snaps, ddi_aging_snaps = read_ddi()

    # DDA datasets.
    dda30_rows, dda_aging_rows, dda30_snaps, dda_aging_snaps = read_dda()

    names = read_names()

    # Ordered dimension lists.
    def ordered(values, preferred):
        out = [v for v in preferred if v in values]
        rest = sorted(v for v in values if v not in preferred and v != "")
        return out + rest

    # Freqs: known order, then any extras; UNSET last if present.
    freq_list = [f for f in FREQ_DISPLAY_ORDER if f in freqs]
    freq_list += sorted(f for f in freqs if f not in FREQ_DISPLAY_ORDER and f != UNSET)
    if UNSET in freqs:
        freq_list.append(UNSET)

    snapshots_sorted = sorted(snapshots)

    data = {
        "generated": "",  # filled by caller stamp (kept blank in-script; no Date in JS runtime)
        "snapshots": snapshots_sorted,
        "latest": snapshots_sorted[-1] if snapshots_sorted else "",
        "months": months_sorted[-KEEP_MONTHS:],
        "schemes": sorted(schemes),
        "trustees": sorted(trustees),
        "statuses": ordered(statuses, STATUS_ORDER),
        "modes": ordered(modes, MODE_ORDER),
        "freqs": freq_list,
        "acctTypes": sorted(acct_types),
        "channels": CHANNELS,
        "rows": rows,
        "pym": pym,
        "aoAging": {"snapshots": ao_aging_snaps, "rows": ao_aging_rows},
        "ddi30": {"snapshots": ddi30_snaps, "rows": ddi30_rows},
        "ddiAging": {"snapshots": ddi_aging_snaps, "rows": ddi_aging_rows},
        "dda30": {"snapshots": dda30_snaps, "rows": dda30_rows},
        "ddaAging": {"snapshots": dda_aging_snaps, "rows": dda_aging_rows},
        "names": names,
    }

    js = "/* AUTO-GENERATED by scripts/build_data.py — do not edit. */\n"
    js += "/* Source: con-bill-6mon (SQL-05) + con-pym-6mon (SQL-06) + con-pym-ao-aging (SQL-07) + ddi-* (SQL-01/02) + dda-* (SQL-03/04) + constant-scheme-info.xlsx */\n"
    js += "const DATA = " + json.dumps(data, separators=(",", ":")) + ";\n"
    with open(OUT, "w", encoding="utf-8") as fh:
        fh.write(js)

    print(f"[build_data] {len(paths)} snapshot(s) -> {OUT}")
    print(f"  snapshots : {data['snapshots']}  (latest {data['latest']})")
    print(f"  months    : {data['months']}  (kept latest {KEEP_MONTHS})")
    print(f"  rows      : {len(rows)}  (of {len(raw_rows)} parsed)")
    print(f"  schemes   : {len(data['schemes'])}  trustees: {len(data['trustees'])}")
    print(f"  statuses  : {len(data['statuses'])}  modes: {len(data['modes'])}  "
          f"freqs: {len(data['freqs'])}  acctTypes: {len(data['acctTypes'])}")
    print(f"  pym rows  : {len(pym_rows)}  snapshots: {pym['snapshots']}")
    print(f"  aoAging   : {len(ao_aging_rows)}  snapshots: {ao_aging_snaps}")
    print(f"  ddi30     : {len(ddi30_rows)}  snapshots: {ddi30_snaps}")
    print(f"  ddiAging  : {len(ddi_aging_rows)}  snapshots: {ddi_aging_snaps}")
    print(f"  dda30     : {len(dda30_rows)}  snapshots: {dda30_snaps}")
    print(f"  ddaAging  : {len(dda_aging_rows)}  snapshots: {dda_aging_snaps}")
    print(f"  names     : {len(names['scheme'])} schemes, {len(names['trustee'])} trustees")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
